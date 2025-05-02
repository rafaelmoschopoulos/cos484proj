from openpyxl import load_workbook

# define your 5-row comparison blocks
BLOCKS = [
    (5, 9),    # block 0: rows 5–9, cols C–W
    (15, 19),  # block 1: rows 15–19, cols C–W
    (25, 29),  # block 2: rows 25–29, cols C–W
    (35, 39),  # block 3: rows 35–39, cols C–H
]

# rows that should be output but NOT used in any comparison
SINGLES = {10, 20, 30}


def compute_stats(ws):
    """
    Build a dict mapping (block_index, col) -> (max, second)
    """
    stats = {}
    for bi, (r0, r1) in enumerate(BLOCKS):
        # columns C–W (3–23) except for block 3 which is C–H (3–8)
        cols = range(3, 9) if bi == 3 else range(3, 24)
        for c in cols:
            vals = [
                ws.cell(row=r, column=c).value
                for r in range(r0, r1+1)
                if isinstance(ws.cell(row=r, column=c).value, (int, float))
            ]
            if not vals:
                continue
            sd = sorted(vals, reverse=True)
            mx = sd[0]
            sec = sd[1] if len(sd) > 1 else None
            stats[(bi, c)] = (mx, sec)
    return stats


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return f"{v:.2f}"
    return str(v)


def format_cell(ws, row, col, stats):
    """
    Return formatted string for cell:
      - column B (2) always text
      - rows in SINGLES just fmt
      - in blocks, bold max, asterisk second
    """
    v = ws.cell(row=row, column=col).value
    # column B: just text
    if col == 2:
        return fmt(v)
    # if it's a single row, just format
    if row in SINGLES:
        return fmt(v)
    # otherwise find which block it's in
    for bi, (r0, r1) in enumerate(BLOCKS):
        if r0 <= row <= r1:
            key = (bi, col)
            if key in stats and isinstance(v, (int, float)):
                mx, sec = stats[key]
                if v == mx:
                    return r"\textbf{" + fmt(v) + "}"
                if sec is not None and v == sec:
                    return fmt(v) + "*"
            break
    return fmt(v)


def save_rows(ws, rows, col_start, col_end, filename, stats):
    """
    Write rows to file from col_start to col_end inclusive.
    """
    lines = []
    for r in rows:
        parts = [
            format_cell(ws, r, c, stats)
            for c in range(col_start, col_end+1)
        ]
        # Use a proper raw string for LaTeX linebreak
        lines.append(" & ".join(parts) + r" \\")
    with open(filename, "w") as f:
        f.write("\n".join(lines))
    print(f"Saved {len(lines)} rows to {filename!r}")

if __name__ == "__main__":
    wb = load_workbook("./tables.xlsx", data_only=True)
    ws = wb["main"]
    stats = compute_stats(ws)

    # include B (2) through W (23) for main, top_p, top_k
    save_rows(ws, range(5, 11),  2, 23, "tab_main.txt",  stats)
    save_rows(ws, range(15,21),  2, 23, "tab_top_p.txt", stats)
    save_rows(ws, range(25,31),  2, 23, "tab_top_k.txt", stats)
    # summary block only B (2) through H (8)
    save_rows(ws, range(35,40),  2,  8, "tab_summary.txt", stats)

