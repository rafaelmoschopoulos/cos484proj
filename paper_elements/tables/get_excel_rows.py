from openpyxl import load_workbook

def format_cell(cell):
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return f"{val:.2f}"
    return str(val)

def excel_rows_to_latex(ws, rows, start_col, end_col):
    """
    Returns a list of LaTeX-ready strings for the given rows & columns in ws.
    """
    lines = []
    for r in rows:
        cells = [ws.cell(row=r, column=c) for c in range(start_col, end_col + 1)]
        formatted = [format_cell(cell) for cell in cells]
        lines.append(" & ".join(formatted) + r" \\")
    return lines

def save_rows(ws, rows, col_start, col_end, filename):
    """
    Writes the LaTeX lines for the specified rows/cols in ws to `filename`.
    """
    lines = excel_rows_to_latex(ws, rows, col_start, col_end)
    with open(filename, "w") as f:
        for line in lines:
            f.write(line + "\n")
    print(f"Saved {len(lines)} rows to {filename!r}")

if __name__ == "__main__":
    path      = "./tables.xlsx"
    sheetname = "main"
    col_start = 2   # B
    col_end   = 23

    wb = load_workbook(path, data_only=True)
    ws = wb[sheetname]

    # each call writes to a separate file
    save_rows(ws, range(5, 11),  col_start, col_end, "tab_main.txt")
    save_rows(ws, range(15, 21), col_start, col_end, "tab_top_p.txt")
    save_rows(ws, range(25, 31), col_start, col_end, "tab_top_k.txt")
    save_rows(ws, range(35, 40), col_start, 8, "tab_summary.txt")
